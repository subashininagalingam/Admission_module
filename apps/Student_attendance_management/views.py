from itertools import count
from multiprocessing import context
# from tkinter.font import Font
from urllib import request, response
from apps.Student_attendance_management.forms import BatchForm

from rest_framework import viewsets
from django.shortcuts import render
from apps.admissions.models import Enrollment
from django.utils import timezone
from rest_framework import generics
from rest_framework import status

from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.utils.timezone import localdate
from rest_framework import filters

from django.db import transaction
from rest_framework import status
from rest_framework.views import APIView

from .filters import AttendanceFilter

from django.http import HttpResponse
import csv

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.platypus import SimpleDocTemplate, Table, Spacer, Paragraph
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet

from django.http import JsonResponse
from django.db.models import Count
from django.utils import timezone

from django.http import JsonResponse
from django.db.models import Q
from django.utils import timezone
from datetime import timedelta

from apps.admissions.models import Course



from .models import (
    Trainer,
    Batch,
    Attendance,
    SyllabusLog
)

from .serializers import (
    TrainerSerializer,
    BatchSerializer,
    AttendanceSerializer,
    SyllabusLogSerializer
)


class TrainerViewSet(viewsets.ModelViewSet):

    queryset = Trainer.objects.all()

    serializer_class = TrainerSerializer


class BatchViewSet(viewsets.ModelViewSet):

    queryset = Batch.objects.all()

    serializer_class = BatchSerializer

    filter_backends = [filters.SearchFilter]

    search_fields = [
        'batch_name',
        'timing',
        'trainer__trainer_name',
        'course__course_name'
    ]

def get_batches_by_course(request):
    course_id = request.GET.get('course_id')

    batches = Batch.objects.filter(course_id=course_id)

    data = [
        {
        "id": b.id,
        "batch_name": b.batch_name
        }
        for b in batches
        if b.student_count < 30
    ]

    return JsonResponse(data, safe=False)

class AttendanceViewSet(viewsets.ModelViewSet):

    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer

    filter_backends = [filters.SearchFilter]

    search_fields = [
        'enrollment__student_first_name',
        'enrollment__student_last_name',
        'enrollment__student_id',
    ]

    def create(self, request, *args, **kwargs):

        enrollment = request.data.get('enrollment')
        batch = request.data.get('batch')
        attendance_status = request.data.get('status')

        attendance_date = timezone.now().date()

        batch_obj = Batch.objects.get(id=batch)

        attendance, created = Attendance.objects.update_or_create(

            enrollment_id=enrollment,
            batch_id=batch,
            attendance_date=attendance_date,

            defaults={
                'status': attendance_status,
                'trainer': batch_obj.trainer
            }
        )

        serializer = self.get_serializer(attendance)

        return Response(
            serializer.data,
            status=status.HTTP_200_OK
        )
    
class AttendanceSubmitAPIView(APIView):

    @transaction.atomic
    def post(self, request):

        batch_id = request.data.get("batch")

        attendance_list = request.data.get(
            "attendance",
            []
        )

        syllabus_data = request.data.get(
            "syllabus_log",
            {}
        )

        try:

            batch = Batch.objects.get(
                id=batch_id
            )

        except Batch.DoesNotExist:

            return Response(
                {
                    "status": False,
                    "message": "Batch not found"
                },
                status=status.HTTP_404_NOT_FOUND
            )

        for item in attendance_list:

            Attendance.objects.update_or_create(

                enrollment_id=item["enrollment"],

                batch=batch,

                attendance_date=timezone.now().date(),

                defaults={
                    "status": item["status"],
                    "remarks": item.get(
                        "remarks",
                        ""
                    ),
                    "trainer": batch.trainer
                }
            )

        SyllabusLog.objects.create(

            batch=batch,

            trainer=batch.trainer,

            topic_covered=syllabus_data[
                "topic_covered"
            ],

            duration=syllabus_data[
                "duration"
            ],

            next_topic=syllabus_data.get(
                "next_topic"
            ),

            trainer_notes=syllabus_data.get(
                "trainer_notes"
            )
        )

        return Response(
            {
                "status": True,
                "message":
                "Attendance and syllabus log saved successfully"
            },
            status=status.HTTP_200_OK
        )
    
class SyllabusLogViewSet(viewsets.ModelViewSet):
    queryset = SyllabusLog.objects.all().order_by('-date')
    serializer_class = SyllabusLogSerializer


def dashboard_api(request):

    search = request.GET.get("search", "")
    today = timezone.now().date()

    batches = Batch.objects.all()

    if search:
        batches = batches.filter(
            Q(batch_name__icontains=search) |
            Q(course__course_name__icontains=search) |
            Q(trainer__trainer_name__icontains=search) |
            Q(timing__icontains=search)
        )

    enrollments = Enrollment.objects.filter(batch__in=batches)

    attendance_qs = Attendance.objects.filter(
        enrollment__in=enrollments,
        attendance_date=today
    )

    return JsonResponse({
        "total": enrollments.count(),
        "present": attendance_qs.filter(status="Present").count(),
        "absent": attendance_qs.filter(status="Absent").count(),
        "late": attendance_qs.filter(status="Late").count(),
    })

def dashboard(request):

    search = request.GET.get("search", "")
    today = timezone.now().date()
    yesterday = today - timedelta(days=1)

    batches = Batch.objects.all()

    if search:
        batches = batches.filter(
            Q(batch_name__icontains=search) |
            Q(course__course_name__icontains=search) |
            Q(trainer__trainer_name__icontains=search) |
            Q(timing__icontains=search)
        )

    enrollments = Enrollment.objects.filter(
        batch__in=batches
    )

    total= enrollments.count()

    attendance_qs = Attendance.objects.filter(
        enrollment__in=enrollments,
        attendance_date=today
    )

  

    today = timezone.now().date()

    # Current Month
    current_month_students = Enrollment.objects.filter(
    start_date__year=today.year,
    start_date__month=today.month
    ).count()

    # Previous Month
    if today.month == 1:
        prev_month = 12
        prev_year = today.year - 1
    else:
        prev_month = today.month - 1
        prev_year = today.year

    previous_month_students = Enrollment.objects.filter(
    start_date__year=prev_year,
    start_date__month=prev_month
    ).count()

    # Percentage
    if previous_month_students > 0:
        total_percentage = round(
            ((current_month_students - previous_month_students) / previous_month_students) * 100, 2
        )
    else:
        total_percentage = 100 if current_month_students > 0 else 0

    present = attendance_qs.filter(status="Present").count()
    absent = attendance_qs.filter(status="Absent").count()
    late = attendance_qs.filter(status="Late").count()

    attendance_marked = attendance_qs.count()

    yesterday_qs = Attendance.objects.filter(
    enrollment__in=enrollments,
    attendance_date=yesterday
    )

    yesterday_present = yesterday_qs.filter(status="Present").count()
    yesterday_absent = yesterday_qs.filter(status="Absent").count()
    yesterday_late = yesterday_qs.filter(status="Late").count()

    percentage = round((present / total) * 100, 2) if total else 0

    present_change = 0
    absent_change = 0
    late_change = 0

    if yesterday_present:
        present_change = round(
            ((present - yesterday_present) / yesterday_present) * 100, 2
        )

    if yesterday_absent:
        absent_change = round(
            ((absent - yesterday_absent) / yesterday_absent) * 100, 2
        )

    if yesterday_late:
        late_change = round(
            ((late - yesterday_late) / yesterday_late) * 100, 2
        )

    context = {
        "present": present,
        "absent": absent,
        "late": late,
        "total": total,
        "percentage": percentage,
        "attendance_marked": attendance_marked,
        "present_change": present_change,
        "absent_change": absent_change,
        "late_change": late_change,
        "total_percentage":total_percentage,
    }

    return render(
        request,
        'attendance/dashboard.html',
        context
    )


def batches_page(request):

    form = BatchForm()

    context = {
        'form': form
    }

    return render(
        request,
        'attendance/batches.html',
        context
    )


def mark_attendance_page(request, batch_id):

    batch = Batch.objects.get(
        id=batch_id
    )

    enrollments = Enrollment.objects.filter(
        batch=batch,
        admission__course_name=batch.course,
    )

    attendance_records = Attendance.objects.filter(
        batch=batch,
        attendance_date=timezone.now().date()
    )

    syllabus_log = SyllabusLog.objects.filter(
    batch=batch,
    date=timezone.now().date()
    ).first()

    duration = syllabus_log.duration if syllabus_log else 0

    attendance_map = {
        att.enrollment_id: att.status
        for att in attendance_records
    }

    remarks_map = {
        att.enrollment_id: att.remarks
        for att in attendance_records
    }

    context = {
        'batch': batch,
        'enrollments': enrollments,
        'attendance_map': attendance_map,
        'remarks_map': remarks_map,
        'syllabus_log': syllabus_log,
        "duration_hours": duration // 60,
        "duration_minutes": duration % 60,
    }

    return render(
        request,
        'attendance/mark_attendance.html',
        context
    )


@api_view(['GET'])
def today_attendance_summary(request, batch_id):

    today = timezone.now().date()

    qs = Attendance.objects.filter(
        batch_id=int(batch_id),  
        attendance_date=today
    )

    data = {
        "present": qs.filter(status="Present").count(),
        "absent": qs.filter(status="Absent").count(),
        "late": qs.filter(status="Late").count(),
        "total": qs.count()
    }

    return Response(data)


@api_view(['POST'])
@transaction.atomic
def bulk_attendance(request):

    print("REQUEST DATA:")
    print(request.data)

    batch_id = request.data.get('batch')

    attendance_list = request.data.get(
        'attendance',
        []
    )

    syllabus_data = request.data.get(
        'syllabus_log',
        {}
    )

    if not syllabus_data.get('topic_covered'):
        return Response(
            {
                'status': False,
                'message': 'Topic covered is required'
            },
            status=400
        )

    duration = syllabus_data.get('duration')

    if not duration:
        return Response(
            {
                'status': False,
                'message': 'Duration is required'
            },
            status=400
        )

    try:

        batch_obj = Batch.objects.get(
            id=batch_id
        )

    except Batch.DoesNotExist:

        return Response(
            {
                'status': False,
                'message': 'Batch not found'
            },
            status=404
        )

    attendance_date = timezone.now().date()

    for item in attendance_list:

        Attendance.objects.update_or_create(

            enrollment_id=item['enrollment'],

            batch_id=batch_id,

            attendance_date=attendance_date,

            defaults={
                'status': item['status'],
                'remarks': item.get(
                    'remarks',
                    ''
                ),
                'trainer': batch_obj.trainer
            }
        )

    SyllabusLog.objects.update_or_create(

    batch=batch_obj,

    date=attendance_date,

    defaults={
        'trainer': batch_obj.trainer,
        'topic_covered': syllabus_data.get(
            'topic_covered'
        ),
        'duration': syllabus_data.get(
            'duration'
        ),
        'next_topic': syllabus_data.get(
            'next_topic',
            ''
        ),
        'trainer_notes': syllabus_data.get(
            'trainer_notes',
            ''
        )
    }
)

    return Response({
        'status': True,
        'message':
        'Attendance and syllabus log saved successfully'
    })

def attendance_report_page(request):

    records = Attendance.objects.select_related(
        'enrollment',
        'batch'
    ).order_by('-attendance_date')
    
    courses = Course.objects.all()

    attendance_filter = AttendanceFilter(
        request.GET,
        queryset=records
    )

    print("GET:", request.GET)
    print("COUNT:", attendance_filter.qs.count())

    context = {
        "filter": attendance_filter,
        "records": attendance_filter.qs,
        "courses": courses
    }

    return render(
        request,
        'attendance/attendance_report.html',
        context
    )


def attendance_export(request):

    records = Attendance.objects.select_related(
    'enrollment',
    'batch',
    'trainer'
).order_by('-attendance_date')

    attendance_filter = AttendanceFilter(request.GET, queryset=records)
    qs = attendance_filter.qs

    export_format = request.GET.get("format")

    # ================= EXCEL =================
    if export_format == "excel":

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        headers = ['Date', 'Student', 'Course', 'Batch', 'Status', 'Trainer']
        ws.append(headers)

        # 🪔 Temple Gold Header Style
        header_fill = PatternFill(
            start_color="FFC000",
            end_color="FFC000",
            fill_type="solid"
        )

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r in qs:
            student_name = f"{r.enrollment.student.first_name} {r.enrollment.student.last_name}"
            course_name = r.enrollment.course.course_name
            batch_name = r.batch.batch_name
            trainer_name = getattr(r.trainer, "name", str(r.trainer))

            ws.append([
                str(r.attendance_date),
                student_name,
                course_name,
                batch_name,
                str(r.status),
                trainer_name
            ])

        # 📏 Column Widths (IMPORTANT: moved OUTSIDE loop)
        column_widths = {
            'A': 20,
            'B': 25,
            'C': 20,
            'D': 18,
            'E': 15,
            'F': 20,
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="attendance_report.xlsx"'

        wb.save(response)
        return response

    # ================= PDF =================
    elif export_format == "pdf":

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="attendance_report.pdf"'

        doc = SimpleDocTemplate(response, pagesize=A4)

        styles = getSampleStyleSheet()
        title = Paragraph("🪔 Attendance Report", styles['Title'])

        data = [['Date', 'Student', 'Course', 'Batch', 'Status', 'Trainer']]

        for r in qs:
            student_name = f"{r.enrollment.student.first_name} {r.enrollment.student.last_name}"
            course_name = r.enrollment.course.course_name
            batch_name = r.batch.batch_name
            trainer_name = getattr(r.trainer, "name", str(r.trainer))

            data.append([
                str(r.attendance_date),
                student_name,
                course_name,
                batch_name,
                str(r.status),
                trainer_name
            ])

        table = Table(data)

        table.setStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.gold),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),

            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),

            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),

            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ])

        elements = [
            title,
            Spacer(1, 12),
            table
        ]

        doc.build(elements)
        return response

    return HttpResponse("Invalid format", status=400)


def student_attendance_summary(request, student_id):

    records = Attendance.objects.filter(
        enrollment__admission__student_id=student_id
    ).order_by("-attendance_date")

    present = records.filter(status="Present").count()
    absent = records.filter(status="Absent").count()
    late = records.filter(status="Late").count()

    total = records.count()

    percentage = round((present / total) * 100, 2) if total else 0

    student = records.first().enrollment.student if records.exists() else None

    timeline = []

    first_record = records.first()

    for r in records:
        timeline.append({
            "date": r.attendance_date.strftime("%d-%m-%Y"),
            "status": r.status
        })

    return JsonResponse({
        "student_name": (
            f"{student.first_name} {student.last_name}"
            if student else ""
        ),
        "student_id": f"STU{student.id}",
        "course": first_record.enrollment.course.course_name if first_record else "",
        "batch": first_record.batch.batch_name if first_record else "",
        "timing": first_record.batch.timing if first_record else "",
        "photo_url": student.photo.url if student and student.photo else None,
        "present": present,
        "absent": absent,
        "late": late,
        "percentage": percentage,
        "timeline": timeline
    })


            # absent tracker

from .services import get_absent_tracker_data


def absent_tracker(request):

    absent_students = get_absent_tracker_data()

    absent_students = [
        student
        for student in absent_students
        if student['total_absences'] > 0
    ]

    absent_students = sorted(
        absent_students,
        key=lambda x: (
            x['consecutive_absences'],
            x['total_absences']
        ),
        reverse=True
    )
    
    courses = Course.objects.all()
    batches = Batch.objects.all()

    return render(
        request,
        'attendance/absent_tracker.html',
        {
            'absent_students': absent_students,
            'courses': courses,
            'batches': batches,
        }
    )
    
from .services import get_low_attendance_data
from django.core.mail import send_mail
from django.conf import settings
from django.contrib import messages
from django.shortcuts import get_object_or_404

def low_attendance_alerts(request):

    low_attendance_students = get_low_attendance_data()

    critical_students = [
        student
        for student in low_attendance_students
        if student["alert_level"] == "Critical"
    ]

    warning_students = [
        student
        for student in low_attendance_students
        if student["alert_level"] == "Warning"
    ]

    return render(
        request,
        'attendance/low_attendance.html',
        {
            'critical_students': critical_students,
            'warning_students': warning_students,
        }
    )


def send_low_attendance_email(request, enrollment_id):

    enrollment = get_object_or_404(
        Enrollment,
        id=enrollment_id
    )

    student = enrollment.student

    attendance_records = Attendance.objects.filter(
        enrollment=enrollment
    )

    total_working_days = Attendance.objects.values(
        'attendance_date'
    ).distinct().count()

    present_count = attendance_records.filter(
        status='Present'
    ).count()

    attendance_percentage = (
        round(
            (present_count / total_working_days) * 100,
            1
        )
        if total_working_days > 0
        else 100
    )

    if attendance_percentage < 60:

        subject = "Critical Attendance Alert"

        message = f"""
Dear {student.first_name},

Your attendance percentage is critically low.

Attendance Percentage: {attendance_percentage}%

Immediate action is required.

Please contact your trainer.

Regards,
CSC Computer Education
"""

    else:

        subject = "Low Attendance Warning"

        message = f"""
Dear {student.first_name},

Your attendance percentage is below the required level.

Attendance Percentage: {attendance_percentage}%

Please attend classes regularly.

Regards,
CSC Computer Education
"""

    send_mail(

        subject,

        message,

        settings.DEFAULT_FROM_EMAIL,

        [student.email],

        fail_silently=False

    )

    messages.success(
        request,
        f"Email sent to {student.email}"
    )

    return JsonResponse({
    "message": "Email sent successfully"
})

def send_sms_notification(request, enrollment_id):

    enrollment = get_object_or_404(
        Enrollment,
        id=enrollment_id
    )

    student = enrollment.student

    attendance_records = Attendance.objects.filter(
        enrollment=enrollment
    )

    total_working_days = Attendance.objects.values(
        'attendance_date'
    ).distinct().count()

    present_count = attendance_records.filter(
        status='Present'
    ).count()

    attendance_percentage = (
        round(
            (present_count / total_working_days) * 100,
            1
        )
        if total_working_days > 0
        else 100
    )

    if attendance_percentage < 60:

        sms_message = f"""
CSC ALERT

Critical Attendance Alert

Student:
{student.first_name}

Attendance:
{attendance_percentage}%

Immediate action required.
"""

    else:

        sms_message = f"""
CSC ALERT

Low Attendance Warning

Student:
{student.first_name}

Attendance:
{attendance_percentage}%

Please attend classes regularly.
"""

    print(sms_message)

    messages.success(
        request,
        f"SMS sent to {student.phone_no}"
    )

    return JsonResponse({
    "message": "SMS sent successfully"
})

def send_email_all(request):

    low_attendance_students = (
        get_low_attendance_data()
    )

    for student_data in low_attendance_students:

        student = student_data["student"]

        attendance_percentage = (
            student_data["attendance_percentage"]
        )

        if attendance_percentage < 60:

            subject = (
                "Critical Attendance Alert"
            )

            message = f"""
Dear {student.first_name},

Your attendance percentage is critically low.

Attendance:
{attendance_percentage}%

Immediate action required.

Regards,
CSC Computer Education
"""

        else:

            subject = (
                "Low Attendance Warning"
            )

            message = f"""
Dear {student.first_name},

Your attendance percentage is below the required level.

Attendance:
{attendance_percentage}%

Please improve attendance.

Regards,
CSC Computer Education
"""

        if student.email:

            send_mail(

                subject,

                message,

                settings.DEFAULT_FROM_EMAIL,

                [student.email],

                fail_silently=False

            )

    return JsonResponse({

        "message":
        "All Emails sent successfully"

    })

def send_sms_all(request):

    low_attendance_students = (
        get_low_attendance_data()
    )

    for student_data in low_attendance_students:

        student = student_data["student"]

        attendance_percentage = (
            student_data["attendance_percentage"]
        )

        if attendance_percentage < 60:

            sms_message = f"""
CSC ALERT

Critical Attendance Alert

Student:
{student.first_name}

Attendance:
{attendance_percentage}%
"""

        else:

            sms_message = f"""
CSC ALERT

Low Attendance Warning

Student:
{student.first_name}

Attendance:
{attendance_percentage}%
"""

        print(sms_message)

    messages.success(
        request,
        "SMS notifications sent."
    )

    return JsonResponse({
    "message": " All SMS notifications sent"
})

def send_monthly_report(request):

    low_attendance_students = (
        get_low_attendance_data()
    )

    report_lines = []

    report_lines.append(
        "Monthly Low Attendance Report\n"
    )

    for student_data in low_attendance_students:

        report_lines.append(

            f"""
Student:
{student_data['student'].first_name}
{student_data['student'].last_name}

Course:
{student_data['course'].course_name}

Batch:
{student_data['batch'].batch_name}

Attendance:
{student_data['attendance_percentage']}%

Total Absences:
{student_data['total_absences']}
"""
        )

    report_content = "\n".join(
        report_lines
    )

    send_mail(

        "Monthly Attendance Report",

        report_content,

        settings.DEFAULT_FROM_EMAIL,

        [settings.DEFAULT_FROM_EMAIL],

        fail_silently=False

    )

    messages.success(
        request,
        "Monthly report sent."
    )

    return JsonResponse({
    "message": "Monthly report sent"
})
    #Reports 
    
from django.shortcuts import render
from django.utils import timezone
from django.db.models import Count

from apps.admissions.models import Enrollment, Course
from .models import Attendance, Batch
from .services import get_low_attendance_data

def get_report_students():

    report_students = []

    enrollments = Enrollment.objects.select_related(
        'admission__student',
        'admission__course_name',
        'batch'
    )

    total_days = Attendance.objects.values(
        'attendance_date'
    ).distinct().count()

    for enrollment in enrollments:

        present_count = Attendance.objects.filter(
            enrollment=enrollment,
            status='Present'
        ).count()

        absent_count = Attendance.objects.filter(
            enrollment=enrollment,
            status='Absent'
        ).count()

        late_count = Attendance.objects.filter(
            enrollment=enrollment,
            status='Late'
        ).count()

        attendance_rate = round(
            (present_count / total_days) * 100,
            1
        ) if total_days > 0 else 0

        if attendance_rate >= 75:
            status = "Good"

        elif attendance_rate >= 60:
            status = "Warning"

        else:
            status = "Critical"

        report_students.append({

            "student": enrollment.student,

            "course": enrollment.course,

            "batch": enrollment.batch,

            "present_count": present_count,

            "absent_count": absent_count,

            "late_count": late_count,

            "attendance_rate": attendance_rate,

            "status": status,

            "total_days": total_days,

        })

    return report_students


def reports(request):

    today = timezone.now().date()
    
    total_students = Enrollment.objects.count()
    
    today_marked_count = Attendance.objects.filter(
        attendance_date=today
    ).count()

    pending_count = total_students - today_marked_count

    if today_marked_count == 0:
 
       attendance_status = "not_started"

    elif pending_count > 0:

        attendance_status = "in_progress"

    else:

         attendance_status = "completed"
    
    

    # Top cards

    

    present_today = Attendance.objects.filter(
        attendance_date=today,
        status='Present'
    ).count()

    absent_today = Attendance.objects.filter(
        attendance_date=today,
        status='Absent'
    ).count()

    low_attendance = len(
        get_low_attendance_data()
    )

    # Report table

    report_students = []

    enrollments = Enrollment.objects.select_related(
        'admission__student',
        'admission__course_name',
        'batch'
    )

    total_days = Attendance.objects.values(
        'attendance_date'
    ).distinct().count()

    for enrollment in enrollments:

        present_count = Attendance.objects.filter(
            enrollment=enrollment,
            status='Present'
        ).count()

        absent_count = Attendance.objects.filter(
            enrollment=enrollment,
            status='Absent'
        ).count()

        late_count = Attendance.objects.filter(
            enrollment=enrollment,
            status='Late'
        ).count()

        attendance_rate = (
            round(
                (present_count / total_days) * 100,
                1
            )
            if total_days > 0 else 0
        )

        if attendance_rate >= 75:
            status = "Good"

        elif attendance_rate >= 60:
            status = "Warning"

        else:
            status = "Critical"

        report_students.append({

            "student":
            enrollment.student,

            "course":
            enrollment.course,

            "batch":
            enrollment.batch,

            "present_count":
            present_count,

            "absent_count":
            absent_count,

            "late_count":
            late_count,

            "attendance_rate":
            attendance_rate,

            "status":
            status,
            
            "total_days": total_days,

        })

    # Monthly Chart

    monthly_present = []
    monthly_absent = []
    monthly_late = []

    for month in range(1, 13):

        monthly_present.append(

            Attendance.objects.filter(
                attendance_date__month=month,
                status='Present'
            ).count()

        )

        monthly_absent.append(

            Attendance.objects.filter(
                attendance_date__month=month,
                status='Absent'
            ).count()

        )

        monthly_late.append(

            Attendance.objects.filter(
                attendance_date__month=month,
                status='Late'
            ).count()

        )

    # Course Analytics

    course_labels = []
    course_counts = []

    courses = Course.objects.all()

    for course in courses:

        course_labels.append(
            course.course_name
        )

        course_counts.append(

            Enrollment.objects.filter(
                admission__course_name=course
            ).count()

        )

    # Batch Analytics

    batch_labels = []
    batch_counts = []
    batch_present_counts = []
    batch_performance_labels = []
    batch_performance_counts = []

    batches = Batch.objects.all()

    for batch in batches:

        batch_labels.append(
            batch.batch_name
        )

        batch_counts.append(

            Enrollment.objects.filter(
                batch=batch
            ).count()

        )
        
        
        
        present_count = Attendance.objects.filter(
            batch=batch,
            status="Present"
        ).count()

        total_count = Attendance.objects.filter(
            batch=batch
        ).count()

        percentage = round(
            (present_count / total_count) * 100,
            1
        ) if total_count else 0

        batch_present_counts.append(
            present_count
        )

        batch_performance_labels.append(
            batch.batch_name
        )

        batch_performance_counts.append(
            percentage
        )

    context = {

        "total_students":
        total_students,

        "present_today":
        present_today,

        "absent_today":
        absent_today,

        "low_attendance":
        low_attendance,

        "report_students":
        report_students,

        "monthly_present":
        monthly_present,

        "monthly_absent":
        monthly_absent,

        "monthly_late":
        monthly_late,

        "course_labels":
        course_labels,

        "course_counts":
        course_counts,

        "batch_labels":
        batch_labels,

        "batch_counts":
        batch_counts,
        
        "batches": batches,
        
        "courses": courses,
        
        "batch_present_counts": batch_present_counts,
        
        "batch_performance_labels": batch_performance_labels,
        
        "batch_performance_counts": batch_performance_counts,
        
        "attendance_status": attendance_status,
        
        "today_marked_count": today_marked_count,
        
        "pending_count": pending_count,
        
        
            

    }

    return render(
        request,
        "attendance/reports.html",
        context
    )
    
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak
)

from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import enums


def analytics_pdf(request):

    response = HttpResponse(
        content_type='application/pdf'
    )

    response['Content-Disposition'] = (
        'attachment; filename=analytics_report.pdf'
    )

    doc = SimpleDocTemplate(response)

    styles = getSampleStyleSheet()

    elements = []

    # Title

    elements.append(

        Paragraph(
            "CSC Computer Education",
            styles['Title']
        )

    )

    elements.append(

        Paragraph(
            "Attendance Analytics Report",
            styles['Heading2']
        )

    )

    elements.append(
        Spacer(1, 12)
    )

    # Summary

    today = timezone.now().date()

    total_students = (
        Enrollment.objects.count()
    )

    present_today = Attendance.objects.filter(
        attendance_date=today,
        status='Present'
    ).count()

    absent_today = Attendance.objects.filter(
        attendance_date=today,
        status='Absent'
    ).count()

    low_attendance = len(
        get_low_attendance_data()
    )

    summary_data = [

        ["Metric", "Value"],

        ["Total Students", total_students],

        ["Present Today", present_today],

        ["Absent Today", absent_today],

        ["Low Attendance", low_attendance]

    ]

    summary_table = Table(
        summary_data
    )

    summary_table.setStyle(

        TableStyle([

            ('BACKGROUND',(0,0),(-1,0),colors.lightgrey),

            ('GRID',(0,0),(-1,-1),1,colors.black),

            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold')

        ])

    )

    elements.append(summary_table)

    elements.append(
        Spacer(1,20)
    )

    # Course Analytics

    elements.append(

        Paragraph(
            "Course Analytics",
            styles['Heading2']
        )

    )

    course_data = [
        ["Course", "Students"]
    ]

    for course in Course.objects.all():

        count = Enrollment.objects.filter(
            admission__course_name=course
        ).count()

        course_data.append([
            course.course_name,
            count
        ])

    course_table = Table(
        course_data
    )

    course_table.setStyle(

        TableStyle([

            ('BACKGROUND',(0,0),(-1,0),colors.lightgrey),

            ('GRID',(0,0),(-1,-1),1,colors.black)

        ])

    )

    elements.append(course_table)

    elements.append(
        Spacer(1,20)
    )

    # Batch Analytics

    elements.append(

        Paragraph(
            "Batch Analytics",
            styles['Heading2']
        )

    )

    batch_data = [
        ["Batch", "Students"]
    ]

    for batch in Batch.objects.all():

        count = Enrollment.objects.filter(
            batch=batch
        ).count()

        batch_data.append([
            batch.batch_name,
            count
        ])

    batch_table = Table(
        batch_data
    )

    batch_table.setStyle(

        TableStyle([

            ('BACKGROUND',(0,0),(-1,0),colors.lightgrey),

            ('GRID',(0,0),(-1,-1),1,colors.black)

        ])

    )

    elements.append(batch_table)

    doc.build(elements)

    return response


def analytics_excel(request):

    wb = Workbook()

    # Sheet 1
    ws1 = wb.active
    ws1.title = "Summary"

    total_students = Enrollment.objects.count()

    today = timezone.now().date()

    present_today = Attendance.objects.filter(
        attendance_date=today,
        status='Present'
    ).count()

    absent_today = Attendance.objects.filter(
        attendance_date=today,
        status='Absent'
    ).count()

    low_attendance = len(
        get_low_attendance_data()
    )

    ws1.append(["Metric", "Value"])

    ws1.append(["Total Students", total_students])

    ws1.append(["Present Today", present_today])

    ws1.append(["Absent Today", absent_today])

    ws1.append(["Low Attendance", low_attendance])

    # Sheet 2
    ws2 = wb.create_sheet("Course Analytics")

    ws2.append([
        "Course",
        "Students"
    ])

    for course in Course.objects.all():

        count = Enrollment.objects.filter(
            admission__course_name=course
        ).count()

        ws2.append([
            course.course_name,
            count
        ])

    # Sheet 3
    ws3 = wb.create_sheet("Batch Analytics")

    ws3.append([
        "Batch",
        "Students"
    ])

    for batch in Batch.objects.all():

        count = Enrollment.objects.filter(
            batch=batch
        ).count()

        ws3.append([
            batch.batch_name,
            count
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = (
        'attachment; filename=analytics_report.xlsx'
    )

    wb.save(response)

    return response


def report_pdf(request):

    response = HttpResponse(
        content_type='application/pdf'
    )

    response['Content-Disposition'] = (
        'attachment; filename=attendance_report.pdf'
    )

    doc = SimpleDocTemplate(response)

    styles = getSampleStyleSheet()

    elements = []

    title = Paragraph(
        "CSC Computer Education - Attendance Report",
        styles['Title']
    )

    elements.append(title)

    elements.append(Spacer(1, 12))

    data = [[

        "Student Name",
        "Course",
        "Batch",
        "Present",
        "Absent",
        "Late",
        "Attendance %",
        "Status"

    ]]

    students = get_report_students()

    for student in students:

        data.append([

            f"{student['student'].first_name} {student['student'].last_name}",

            student['course'].course_name,

            student['batch'].batch_name
            if student['batch'] else "-",

            student['present_count'],

            student['absent_count'],

            student['late_count'],

            f"{student['attendance_rate']}%",

            student['status']

        ])

    table = Table(data)

    table.setStyle(

        TableStyle([

            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),

            ('GRID', (0, 0), (-1, -1), 1, colors.black),

            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),

            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),

        ])

    )

    elements.append(table)

    doc.build(elements)

    return response


def report_excel(request):

    wb = Workbook()

    ws = wb.active

    ws.title = "Attendance Report"

    headers = [

        "Student Name",
        "Course",
        "Batch",
        "Present Days",
        "Absent Days",
        "Late Days",
        "Attendance %",
        "Status"

    ]

    ws.append(headers)

    students = get_report_students()

    for student in students:

        ws.append([

            f"{student['student'].first_name} {student['student'].last_name}",

            student['course'].course_name,

            student['batch'].batch_name
            if student['batch'] else "-",

            student['present_count'],

            student['absent_count'],

            student['late_count'],

            student['attendance_rate'],

            student['status']

        ])

    response = HttpResponse(

        content_type=
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    )

    response['Content-Disposition'] = (
        'attachment; filename=attendance_report.xlsx'
    )

    wb.save(response)

    return response